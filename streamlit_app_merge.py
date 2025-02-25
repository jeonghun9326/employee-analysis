import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook

def upload_excel_files():
    """ Streamlit UI에서 다중 엑셀 파일을 업로드하는 함수 """
    return st.file_uploader("📂 엑셀 파일을 선택하세요", type=["xlsx"], accept_multiple_files=True)

def get_delete_keywords():
    """ Streamlit UI에서 키워드 기반 삭제 컬럼을 입력받는 함수 """
    st.sidebar.subheader("🔒 개인정보 보호 설정")

    # 사용자가 쉼표로 구분하여 키워드를 입력하면 리스트로 변환
    delete_keywords_input = st.sidebar.text_area("🔍 키워드로 삭제할 컬럼 입력 (쉼표로 구분)", "연봉")
    delete_keywords = [kw.strip() for kw in delete_keywords_input.split(",") if kw.strip()]
    
    return delete_keywords

def get_include_columns():
    """ Streamlit UI에서 추출할 컬럼을 입력받는 함수 """
    st.sidebar.subheader("🔒 추출할 컬럼 설정")

    # 사용자가 쉼표로 구분하여 추출할 컬럼을 입력하면 리스트로 변환
    include_columns_input = st.sidebar.text_area("🔍 추출할 컬럼 입력 (쉼표로 구분)", "성명, 입사일, 퇴사일")
    include_columns = [col.strip() for col in include_columns_input.split(",") if col.strip()]
    
    return include_columns

def merge_excel_files(uploaded_files, delete_keywords, include_columns):
    """ 업로드된 다수의 엑셀 파일을 하나의 파일로 병합 """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for file in uploaded_files:
            file_name = file.name.split('.')[0]  # 파일명에서 확장자 제거
            xls = pd.ExcelFile(file, engine='openpyxl')  # openpyxl로 엑셀 파일 로드
            
            for sheet_name in xls.sheet_names:
                sheet_df = pd.read_excel(xls, sheet_name=sheet_name, engine='openpyxl')
                
                # 사용자가 지정한 컬럼만 추출
                if include_columns:
                    columns_to_include = [col for col in sheet_df.columns if col in include_columns]
                    sheet_df = sheet_df[columns_to_include]

                # 키워드에 해당하는 컬럼 삭제
                columns_to_delete = [col for col in sheet_df.columns if any(keyword in col for keyword in delete_keywords)]
                sheet_df.drop(columns=columns_to_delete, inplace=True, errors='ignore')

                # 엑셀 파일의 서식을 복사하기 위한 작업
                wb = load_workbook(file)
                sheet = wb[sheet_name]
                
                new_sheet_name = f"{file_name}"  # 파일명_원래시트명 형식
                sheet_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                
                # 시트의 열 너비, 행 높이, 숫자 표기법, 날짜 표기법 복사
                worksheet = writer.sheets[new_sheet_name]
                
                # 열 너비 자동 조정
                for col in sheet.columns:
                    column = col[0].column_letter  # 열 번호 (A, B, C, ...)
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # 여유 공간을 위해 2 추가
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # 행 높이 복사
                for row in sheet.iter_rows():
                    row_height = sheet.row_dimensions[row[0].row].height
                    worksheet.row_dimensions[row[0].row].height = row_height
                
                # 숫자 표기법 및 날짜 표기법 복사
                for row in sheet.iter_rows():
                    for cell in row:
                        new_cell = worksheet[cell.coordinate]
                        if cell.number_format:
                            new_cell.number_format = cell.number_format

    output.seek(0)
    return output

def run_excel_merge():
    """ Streamlit에서 엑셀 병합 기능 실행 """
    st.title("엑셀 파일 병합기")
    st.write("다수의 엑셀 파일을 업로드하여 하나의 파일로 병합합니다. 각 파일의 내용은 파일명과 동일한 시트명으로 저장됩니다.")

    include_columns = get_include_columns()  # 추출할 컬럼만 입력 받기
    delete_keywords = get_delete_keywords()
    uploaded_files = upload_excel_files()
    
    if not uploaded_files:
        st.warning("⚠️ 하나 이상의 엑셀 파일을 업로드해주세요.")
        return
    
    st.success(f"{len(uploaded_files)}개의 파일이 업로드되었습니다.")  # 업로드된 파일 개수 확인
    
    # 병합된 엑셀 파일 생성
    merged_file = merge_excel_files(uploaded_files, delete_keywords, include_columns)

    # 다운로드 버튼 추가
    st.download_button(
        label="📥 병합된 엑셀 파일 다운로드",
        data=merged_file,
        file_name="merged_excel.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

