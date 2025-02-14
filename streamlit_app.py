import os
import pandas as pd
import zipfile
import tempfile
import shutil
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime, timedelta

# 📌 현재 날짜 기준 전월 및 당월 계산
today = datetime.today()
current_month = today.strftime("%Y-%m")
previous_month = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
previous_month_last_day = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")

# 📌 분석 대상 컬럼 설정
date_columns = ["입사일", "퇴사일"]
employee_types = ["정규직", "계약직", "파견직", "임원"]  # 가나다순 정렬

# 📌 시트 정렬 순서
sheet_order = [
    "도이치아우토",
    "브리티시오토",
    "바이에른오토",
    "이탈리아오토모빌리",
    "브리타니아오토",
    "디티네트웍스",
    "도이치파이낸셜",
    "BAMC",
    "차란차",
    "디티이노베이션",
    "도이치오토월드",
    "DAFS",
    "사직오토랜드"
]

# 📌 Streamlit UI
st.title("📊 인원 분석 자동화 시스템 (ZIP 파일 업로드 지원)")
st.write("ZIP 파일을 업로드하면 자동으로 엑셀 병합 및 인원 분석을 수행합니다.")

# 📌 ZIP 파일 업로드
uploaded_zip = st.file_uploader("📂 ZIP 파일을 업로드하세요", type=["zip"])

if uploaded_zip:
    try:
        # 📌 임시 폴더 생성
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, "uploaded.zip")

        # 📌 ZIP 파일 저장
        with open(zip_path, "wb") as f:
            f.write(uploaded_zip.read())

        # 📌 ZIP 압축 해제
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)

        # 📌 병합된 파일 경로 설정
        merged_excel_path = os.path.join(temp_dir, "merged_excel.xlsx")

        # 📌 엑셀 병합 함수 실행
        def merge_excel_files(folder_path, output_file):
            files = [file for file in os.listdir(folder_path) if file.endswith(".xlsx") and not file.startswith("~$")]
            files.sort(key=lambda x: sheet_order.index(os.path.splitext(x)[0]) if os.path.splitext(x)[0] in sheet_order else len(sheet_order))

            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for file in files:
                    file_path = os.path.join(folder_path, file)
                    try:
                        wb = load_workbook(file_path, data_only=True)
                        sheet_names = wb.sheetnames  # 모든 시트 포함 (숨겨진 시트 포함)

                        if not sheet_names:
                            st.warning(f"⚠️ 파일 `{file}` 에 사용 가능한 시트가 없어 건너뜁니다.")
                            continue

                        for sheet_name in sheet_names:
                            ws = wb[sheet_name]
                            data = [[cell.value for cell in row] for row in ws.iter_rows()]
                            
                            if not data or all(all(cell is None for cell in row) for row in data):
                                st.warning(f"⚠️ 파일 `{file}` 의 시트 `{sheet_name}` 가 비어 있어 건너뜁니다.")
                                continue

                            header_row_index = None
                            for idx, row in enumerate(data):
                                if row[0] == "NO":
                                    header_row_index = idx
                                    break

                            if header_row_index is not None:
                                headers = data[header_row_index]
                                df = pd.DataFrame(data[header_row_index + 1:], columns=headers)
                            else:
                                df = pd.DataFrame(data[1:], columns=data[0])

                            sheet_name_trimmed = os.path.splitext(file)[0][:31]
                            df.to_excel(writer, sheet_name=sheet_name_trimmed, index=False)

                    except Exception as e:
                        st.error(f"🚨 파일 `{file}` 처리 중 오류 발생: {e}")

        merge_excel_files(temp_dir, merged_excel_path)
        st.success("✅ 엑셀 파일 병합 완료!")

        # 📌 병합된 엑셀 파일 분석 시작
        sheets = pd.read_excel(merged_excel_path, sheet_name=None, engine="openpyxl")

        for sheet_name, df in sheets.items():
            st.subheader(f"📄 시트 이름: {sheet_name}")

            # "Starting Date" → "입사일"로 변경
            if "Starting Date" in df.columns:
                df.rename(columns={"Starting Date": "입사일"}, inplace=True)

            # 컬럼명 공백 제거
            df.columns = df.columns.str.strip()

            # 미국식 날짜 변환
            if "입사일" in df.columns:
                df["입사일"] = pd.to_datetime(df["입사일"], errors="coerce").dt.strftime("%Y-%m-%d")

            # "퇴사일" 컬럼 생성
            if "퇴사일" not in df.columns:
                df["퇴사일"] = None

            # "Remark" 컬럼 값에 따라 "퇴사일" 설정
            if "Remark" in df.columns:
                df.loc[df["Remark"].astype(str).str.startswith("Resigned and last working"), "퇴사일"] = previous_month_last_day

            # "사원구분명" 컬럼 생성
            if "사원구분명" not in df.columns:
                df["사원구분명"] = None

            # "Contract Type" 기준으로 "사원구분명" 설정
            if "Contract Type" in df.columns:
                df.loc[df["Contract Type"].astype(str).str.contains("FDC", na=False), "사원구분명"] = "비정규직"
                df.loc[df["Contract Type"].astype(str).str.contains("UDC", na=False), "사원구분명"] = "정규직"

            # 날짜 변환
            for col in date_columns:
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m")

            # 결과 출력
            st.write(f"📌 전월({previous_month}) 입사자 수: {df[df['입사일'] == previous_month].shape[0]}")
            st.write(f"📌 전월({previous_month}) 퇴사자 수: {df[df['퇴사일'] == previous_month].shape[0]}")

        st.download_button(label="📥 병합된 엑셀 다운로드", data=open(merged_excel_path, "rb"), file_name="merged_excel.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"❌ 오류 발생: {e}")
    finally:
        shutil.rmtree(temp_dir)  # 임시 폴더 삭제


        st.download_button(label="📥 병합된 엑셀 다운로드", data=open(merged_excel_path, "rb"), file_name="merged_excel.xlsx", mime="application/vnd.o

